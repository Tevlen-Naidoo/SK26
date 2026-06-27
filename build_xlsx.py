#!/usr/bin/env python3
"""Build the consolidated TNT South Korea 2026 itinerary spreadsheet."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- styles ----
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SEOUL_FILL = PatternFill("solid", fgColor="DDEBF7")
JEJU_FILL = PatternFill("solid", fgColor="E2EFDA")
BUSAN_FILL = PatternFill("solid", fgColor="FCE4D6")
DAY_FILL = PatternFill("solid", fgColor="BDD7EE")
DAY_FONT = Font(bold=True, size=11, color="1F4E78")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

RATE = 93  # 1 ZAR = 93 KRW (June 2026)

def gmap(q):
    return "https://www.google.com/maps/search/?api=1&query=" + q.replace(" ", "+")
def kmap(q):
    return "https://map.kakao.com/?q=" + q.replace(" ", "%20")

def cost_cell(raw):
    """raw is a string of digits/commas, '0', or ''. Return dual-currency text."""
    s = str(raw).replace(",", "").strip()
    if s == "":
        return ""
    krw = int(s)
    if krw == 0:
        return "Free"
    zar = round(krw / RATE)
    return f"₩{krw:,} / R{zar:,}"

# ============================================================
# SHEET 1 — Day-by-day itinerary
# ============================================================
ws = wb.active
ws.title = "Itinerary"
headers = ["Date", "Day", "Time", "Activity", "Area / Notes", "Cost pp (₩ / R)", "Google Maps", "Kakao Maps"]
widths = [12, 6, 13, 40, 40, 20, 16, 16]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(1, i, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# rows: (date, day, time, activity, notes, cost, mapquery)  -- mapquery None => no link
# Special markers: ("DAY", title) and ("REGION", title)
rows = [
    ("REGION", "SEOUL — 19–25 July (Hotel: Friendly DH Naissance, Seongbuk-gu)"),

    ("DAY", "Sat 18 Jul — Fly out: Johannesburg -> Dubai"),
    ("18 Jul","Sat","10:00","Check-in OR Tambo (JNB) — Emirates EK762","International departure","",None),
    ("18 Jul","Sat","13:30-23:40","EK762 Johannesburg -> Dubai","Depart 13:30 SAST, arrive 23:40 GST; ~8h","",None),

    ("DAY", "Sun 19 Jul — Dubai -> Seoul, then arrival"),
    ("19 Jul","Sun","00:40","Check-in Dubai (DXB) — Emirates EK322","Short overnight layover","",None),
    ("19 Jul","Sun","03:40-17:00","EK322 Dubai -> Seoul (ICN)","Depart 03:40 GST, land 17:00 KST; ~9h","",None),
    ("19 Jul","Sun","17:00","Land Incheon (ICN) — immigration + bags ~1h","","",None),
    ("19 Jul","Sun","18:00-19:30","AREX/Limousine bus to hotel, check in","~1.5h to hotel","",None),
    ("19 Jul","Sun","20:00-21:30","Relaxed dinner — Daehangno / Seongbuk-dong","Near hotel; cheap student eateries","15,000","Marronnier Park Daehangno Seoul"),
    ("19 Jul","Sun","Optional","Naksan Park city-wall night view","10 min walk, free","0","Naksan Park Seoul"),

    ("DAY", "Mon 20 Jul — Royal Seoul: Gyeongbokgung & Bukchon (palace OPEN Mon)"),
    ("20 Jul","Mon","08:00-09:00","Breakfast at hotel","Free (incl.)","0",None),
    ("20 Jul","Mon","09:30-10:00","Gwanghwamun Sq — King Sejong & Admiral Yi statues","Free, outdoors","0","Statue of King Sejong Gwanghwamun Square"),
    ("20 Jul","Mon","10:00-12:30","Gyeongbokgung Palace + 10:00 Guard Ceremony","Closed Tue; FREE in hanbok","3,000","Gyeongbokgung Palace"),
    ("20 Jul","Mon","12:30-14:00","National Folk Museum of Korea (+ lunch)","Inside palace grounds; open daily","0","National Folk Museum of Korea"),
    ("20 Jul","Mon","14:00-16:30","Bukchon Hanok Village + Samcheong-dong","Tourist hrs Mon-Sat 10-17; residential","0","Bukchon Hanok Village"),
    ("20 Jul","Mon","16:30-18:00","Insadong (crafts, tea, Ssamzigil)","Free to wander","0","Insadong Seoul"),
    ("20 Jul","Mon","18:00-20:00","Dinner at Gwangjang Market","Famous street food, cheap","15,000","Gwangjang Market Seoul"),
    ("20 Jul","Mon","On return","Cheonggyecheon Stream night walk","Free","0","Cheonggyecheon Stream"),

    ("DAY", "Tue 21 Jul — Museums of Yongsan & Hannam (all OPEN Tue, clustered)"),
    ("21 Jul","Tue","08:00-09:00","Breakfast","Free","0",None),
    ("21 Jul","Tue","09:30-13:00","National Museum of Korea (+ lunch)","Huge; free; Tue 09:30-17:30","0","National Museum of Korea"),
    ("21 Jul","Tue","13:30-16:00","War Memorial of Korea","Free; closed Mon; ~1.5km away","0","War Memorial of Korea"),
    ("21 Jul","Tue","16:30-18:00","Leeum Museum of Art (optional)","Closed Mon; perm. free w/ booking","0","Leeum Museum of Art"),
    ("21 Jul","Tue","18:30+","Dinner Itaewon / or N Seoul Tower sunset","International food scene","20,000","Itaewon Seoul"),

    ("DAY", "Wed 22 Jul — Everland Theme Park (full day, fireworks)"),
    ("22 Jul","Wed","07:30","Early breakfast, depart (~1h to Yongin)","Shuttle/subway+bus","0",None),
    ("22 Jul","Wed","10:00-22:00","EVERLAND — T-Express, Panda World, parade, FIREWORKS","Buy on Klook ~39k vs gate ~62k","39,000","Everland Theme Park"),

    ("DAY", "Thu 23 Jul — Seoul Forest -> Seongsu -> COEX/Gangnam -> Banpo Fountain"),
    ("23 Jul","Thu","08:00-09:00","Breakfast","Free","0",None),
    ("23 Jul","Thu","09:30-11:30","Seoul Forest (deer, gardens, trails)","Free, open daily","0","Seoul Forest"),
    ("23 Jul","Thu","11:30-14:00","Seongsu-dong cafe district (+ lunch)","Seoul's 'Brooklyn'","20,000","Seongsu-dong cafe street"),
    ("23 Jul","Thu","14:30-18:30","COEX: Starfield Library, K-pop Sq, Gangnam Style statue, Bongeunsa","All one complex; free entry","0","Starfield Library COEX"),
    ("23 Jul","Thu","18:30-19:30","Dinner in Gangnam","","20,000",None),
    ("23 Jul","Thu","20:00 or 20:30","Banpo Bridge Moonlight Rainbow Fountain","FREE; 20 min; weather-dependent","0","Banpo Bridge Moonlight Rainbow Fountain"),

    ("DAY", "Fri 24 Jul — Seokpajeong & Buam-dong -> MMCA -> Namsan finale"),
    ("24 Jul","Fri","08:00-09:00","Breakfast","Free","0",None),
    ("24 Jul","Fri","10:00-11:30","Seokpajeong garden villa (Buam-dong)","Wed-Sun 10-18; incl. Seoul Museum","20000","Seokpajeong Seoul"),
    ("24 Jul","Fri","11:30-13:30","Buam-dong cafes + city wall (+ lunch)","Sanmotoonge cafe terrace","15,000","Buam-dong Seoul"),
    ("24 Jul","Fri","14:00-16:00","MMCA Seoul (modern art, Samcheong)","Closed Mon; ~10 min from Buam","4,000","MMCA Seoul"),
    ("24 Jul","Fri","16:00-17:30","Seoul Museum of History (optional)","Free; closed Mon; Fri to 21:00","0","Seoul Museum of History"),
    ("24 Jul","Fri","18:00-21:00","N Seoul Tower / Namsan sunset + Myeongdong dinner","Cable car ~14k, obs ~21k","21,000","N Seoul Tower"),

    ("DAY", "Sat 25 Jul — Fly to Jeju"),
    ("25 Jul","Sat","08:00-10:00","Breakfast, pack, check out (by 12:00)","Optional local stroll","0",None),
    ("25 Jul","Sat","~12:00","At Gimpo Airport (GMP)","~50 min from hotel","0","Gimpo Airport"),
    ("25 Jul","Sat","13:55-15:10","TW717 GMP -> Jeju (CJU)","1h15","",None),

    ("REGION", "JEJU — 25–28 July (Hotel: Hotel Moon, Jeju City) — RENT A CAR + bring IDP!"),

    ("DAY", "Sat 25 Jul — Arrival & Jeju City evening"),
    ("25 Jul","Sat","15:10","Land, collect rental car, drive to hotel","~15 min; check in 14:00","",None),
    ("25 Jul","Sat","16:30-18:00","Yongduam (Dragon Head) Rock + coast sunset","Free","0","Yongduam Rock Jeju"),
    ("25 Jul","Sat","18:30-20:30","Dongmun Market — black pork, seafood, hallabong","Dinner","20,000","Dongmun Traditional Market Jeju"),

    ("DAY", "Sun 26 Jul — EAST Jeju: UNESCO sites + Udo Island (full day)"),
    ("26 Jul","Sun","08:30","Breakfast, drive east","","0",None),
    ("26 Jul","Sun","09:30-11:00","Manjanggul Lava Tube (UNESCO)","09-18; closed 1st Wed (not today)","4,000","Manjanggul Lava Tube"),
    ("26 Jul","Sun","11:30-12:30","Seongeup Folk Village (en route)","Historic stone village; free","0","Seongeup Folk Village"),
    ("26 Jul","Sun","12:30-13:30","Lunch near Seongsan (seafood)","","15,000",None),
    ("26 Jul","Sun","13:30-15:00","Seongsan Ilchulbong (Sunrise Peak, UNESCO)","Hike ~40min; closed 1st Mon (not today)","5,000","Seongsan Ilchulbong"),
    ("26 Jul","Sun","15:30-18:30","Udo Island — ferry from Seongsan Port","~15min crossing; check last ferry!","10,500","Udo Island Jeju ferry Seongsan Port"),
    ("26 Jul","Sun","19:30","Drive back / dinner","Alt: Seopjikoji + Aqua Planet","15,000",None),

    ("DAY", "Mon 27 Jul — SOUTH Jeju: Seogwipo, waterfalls & art (full day)"),
    ("27 Jul","Mon","08:30","Breakfast, drive south","","0",None),
    ("27 Jul","Mon","09:30-11:00","Jeju Folk Village (Pyoseon, en route)","Apr-Sep 08:30-18:00","15,000","Jeju Folk Village Pyoseon"),
    ("27 Jul","Mon","11:30-12:30","Jusangjeolli Cliffs (basalt columns)","","2,000","Jusangjeolli Cliff Jeju"),
    ("27 Jul","Mon","12:30-13:30","Lunch Jungmun/Seogwipo","","15,000",None),
    ("27 Jul","Mon","13:30-14:30","Cheonjiyeon Waterfall","Easy paved walk","2,000","Cheonjiyeon Waterfall Seogwipo"),
    ("27 Jul","Mon","15:00-16:30","The Garden of Van Gogh (digital art)","= your 'Van Gogh Garden'; cheaper online","13,000","Garden of Van Gogh Jeju"),
    ("27 Jul","Mon","16:30-18:00","Alive Museum Jeju (Jungmun) — trick art","= your 'Alive Museum'","14,000","Alive Museum Jeju Jungmun"),
    ("27 Jul","Mon","Optional","Seogwipo Submarine / Grimm Forest / O'Sulloc","See notes for swaps","",None),
    ("27 Jul","Mon","Evening","Drive back; Black Pork Street dinner","Heukdwaeji-geori","25,000","Jeju Black Pork Street"),

    ("DAY", "Tue 28 Jul — Fly to Busan"),
    ("28 Jul","Tue","08:00","Breakfast, pack, return rental car","","0",None),
    ("28 Jul","Tue","10:45-11:45","7C504 Jeju -> Busan (PUS)","Be at CJU by 09:15","",None),

    ("REGION", "BUSAN — 28 July–1 Aug (Hotel: Ososo, Dong-gu, near Busan Station)"),

    ("DAY", "Tue 28 Jul — Arrival & Nampo old town"),
    ("28 Jul","Tue","11:45","Land, to hotel (~30-40 min), drop bags","","",None),
    ("28 Jul","Tue","14:00-16:00","Jagalchi Fish Market + BIFF Square","Seafood + street food","20,000","Jagalchi Market Busan"),
    ("28 Jul","Tue","16:00-18:00","Gukje Market + Yongdusan / Busan Tower","Harbour views","0","Gukje Market Busan"),
    ("28 Jul","Tue","18:00","Seafood dinner in Nampo","","25,000",None),

    ("DAY", "Wed 29 Jul — West Busan: Gamcheon & Songdo"),
    ("29 Jul","Wed","09:30-12:00","Gamcheon Culture Village","Free; ~2 stamp map; ~2h","2,000","Gamcheon Culture Village"),
    ("29 Jul","Wed","12:00-13:00","Lunch","","15,000",None),
    ("29 Jul","Wed","13:30-15:30","Songdo Marine Cable Car + Skywalk","RT general; crystal floor ₩22k","17,000","Songdo Marine Cable Car Busan"),
    ("29 Jul","Wed","16:00-18:30","Huinnyeoul Culture Village (Yeongdo) sunset","Optional Taejongdae cliffs","0","Huinnyeoul Culture Village"),

    ("DAY", "Thu 30 Jul — East coast: Ocean temple, Sky Capsule & yacht"),
    ("30 Jul","Thu","09:00-11:00","Haedong Yonggungsa (ocean temple)","Free; go early","0","Haedong Yonggungsa Temple"),
    ("30 Jul","Thu","11:30-13:00","Drive to Haeundae + lunch","","15,000",None),
    ("30 Jul","Thu","13:00-15:00","Haeundae Blueline Park — Sky Capsule + Beach Train","₩35k per 2-seat capsule; BOOK ONLINE","17,500","Haeundae Blueline Park Sky Capsule"),
    ("30 Jul","Thu","15:00-17:30","Haeundae Beach + Dongbaek / Nurimaru","","0","Haeundae Beach"),
    ("30 Jul","Thu","18:00-20:00","The Bay 101 — night yacht + skyline (+ dinner)","= your 'yacht ride'","25,000","The Bay 101 Busan yacht"),

    ("DAY", "Fri 31 Jul — Gwangalli beach day & Baseball (VERIFY schedule)"),
    ("31 Jul","Fri","Day","Gwangalli Beach + cafes (relaxed)","Drone show is Sat-only (n/a)","0","Gwangalli Beach Busan"),
    ("31 Jul","Fri","Evening","Lotte Giants @ Sajik Stadium (VERIFY home dates)","If away, move to 28-30 Jul night","15,000","Sajik Baseball Stadium Busan"),

    ("DAY", "Sat 1 Aug — Busan -> Incheon -> Home (TRAVEL DAY)"),
    ("1 Aug","Sat","08:00-10:30","Breakfast, pack, easy morning near Busan Stn","","0",None),
    ("1 Aug","Sat","by 12:00","Check out, go to Busan Station","","0","Busan Station KTX"),
    ("1 Aug","Sat","~12:30-13:30","KTX Busan -> Seoul (~2h45)","Book in advance","59,800","Busan Station KTX"),
    ("1 Aug","Sat","then","AREX Seoul Stn -> Incheon T1 (~50 min)","Express train; total ~3.5-4h","4,750","Incheon Airport Terminal 1"),
    ("1 Aug","Sat","20:40","Check-in Incheon T1 — Emirates EK323","International departure","",None),
    ("1 Aug","Sat","23:55","EK323 Seoul (ICN) -> Dubai","Depart 23:55 KST, arrive 2 Aug 04:25 GST","",None),

    ("DAY", "Sun 2 Aug — Dubai -> Johannesburg (home)"),
    ("2 Aug","Sun","06:55","Check-in Dubai (DXB) — Emirates EK763","","",None),
    ("2 Aug","Sun","09:55-16:15","EK763 Dubai -> Johannesburg (JNB)","Depart 09:55 GST, arrive 16:15 SAST. TRIP DONE","",None),
]

r = 2
for row in rows:
    if row[0] == "REGION":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, row[1]); c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="left", vertical="center"); c.border = BORDER
        ws.row_dimensions[r].height = 20
        r += 1; continue
    if row[0] == "DAY":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, row[1]); c.fill = DAY_FILL; c.font = DAY_FONT
        c.alignment = Alignment(horizontal="left", vertical="center"); c.border = BORDER
        ws.row_dimensions[r].height = 18
        r += 1; continue
    date, day, time, act, notes, cost, mq = row
    vals = [date, day, time, act, notes, cost_cell(cost),
            ("Open" if mq else ""), ("Open" if mq else "")]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(r, ci, v); c.border = BORDER; c.alignment = WRAP
    if mq:
        g = ws.cell(r, 7); g.value = "Google"; g.hyperlink = gmap(mq); g.font = Font(color="0563C1", underline="single")
        k = ws.cell(r, 8); k.value = "Kakao"; k.hyperlink = kmap(mq); k.font = Font(color="0563C1", underline="single")
    ws.row_dimensions[r].height = 30
    r += 1

# ============================================================
# SHEET 2 — Flights
# ============================================================
wf = wb.create_sheet("Flights")
fh = ["Date","Route","Flight","Depart","Arrive","Check-in / Notes"]
fw = [12, 26, 12, 22, 26, 26]
for i,(h,w) in enumerate(zip(fh,fw),1):
    c = wf.cell(1,i,h); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
    wf.column_dimensions[get_column_letter(i)].width=w
wf.freeze_panes="A2"
flights = [
    ("18 Jul","JNB -> DXB","EK762","OR Tambo 13:30","Dubai 23:40","Check-in 10:00"),
    ("19 Jul","DXB -> ICN","EK322","Dubai 03:40","Seoul ICN 17:00","Check-in 00:40"),
    ("25 Jul","GMP -> CJU","TW717","Gimpo 13:55","Jeju 15:10","1h15 domestic"),
    ("28 Jul","CJU -> PUS","7C504","Jeju 10:45","Busan 11:45","Booking ZGU52K"),
    ("1 Aug","ICN -> DXB","EK323","Incheon T1 23:55","Dubai 04:25 (2 Aug)","Check-in 20:40"),
    ("2 Aug","DXB -> JNB","EK763","Dubai 09:55","OR Tambo 16:15","Check-in 06:55"),
]
for ri,fr in enumerate(flights,2):
    for ci,v in enumerate(fr,1):
        c=wf.cell(ri,ci,v); c.border=BORDER; c.alignment=WRAP

# ============================================================
# SHEET 3 — Key changes & warnings
# ============================================================
ww = wb.create_sheet("Changes & Warnings")
wh=["#","Issue found","Fix / Action"]; wwd=[5,55,55]
for i,(h,w) in enumerate(zip(wh,wwd),1):
    c=ww.cell(1,i,h); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
    ww.column_dimensions[get_column_letter(i)].width=w
ww.freeze_panes="A2"
warns=[
    (1,"20 Jul is a MONDAY. MMCA, War Memorial, Leeum, Seoul Museum of History all closed Mondays.","Monday = Gyeongbokgung day (palace open Mon, closed Tue). Monday-closed museums moved to Tue/Fri."),
    (2,"Original Mon mixed Yongsan (Nat. Museum) with downtown Gwanghwamun ~5km apart.","Nat. Museum + War Memorial + Leeum are all in Yongsan/Hannam -> one day (Tue)."),
    (3,"Arrival-night COEX/Gangnam (18:30) impossible (reach hotel ~19:30, 20+h travel).","Arrival = relaxed dinner near hotel. COEX moved to Thu."),
    (4,"COEX listed twice; K-pop Sq, Starfield, Gangnam Style statue all at SAME complex.","Consolidated into one COEX visit (Thu)."),
    (5,"Gwangalli M Drone Show is SATURDAYS only; only Busan Sat = 1 Aug = departure day.","Drone show NOT feasible. Replaced with The Bay 101 night view + Gwangalli evening."),
    (6,"1 Aug flight departs INCHEON, not Busan. Busan->ICN = 3.5-4h.","1 Aug is a travel day: relaxed Busan morning -> midday KTX -> ICN."),
    (7,"Jeju public transport is poor.","Rent a car (~55-70k/day). MUST get International Driving Permit in SA before departure."),
    (8,"'Submarine to Udo' doesn't exist.","Udo via FERRY (east). Seogwipo Submarine is separate (south, optional)."),
    (9,"Bukchon Hanok Village closed to tourists Sundays; hours Mon-Sat 10-17.","Placed on Monday; be quiet (residential)."),
    (10,"Seokpajeong open Wed-Sun 10-18 only.","Placed Friday 24 Jul."),
    (11,"Museum San is in Wonju, 140 km east of Seoul.","Dropped - too far for this trip."),
    (12,"2026 KBO late-July schedule only in images; not machine-verified.","VERIFY Lotte home dates at koreabaseball.com; slot game into whichever Busan night they're home."),
    (13,"Hanbok tip.","Rent hanbok near Gyeongbokgung -> FREE palace entry while wearing it."),
]
for ri,wr in enumerate(warns,2):
    for ci,v in enumerate(wr,1):
        c=ww.cell(ri,ci,v); c.border=BORDER; c.alignment=WRAP
        if ci>1: c.fill=WARN_FILL
    ww.row_dimensions[ri].height=42

# ============================================================
# SHEET 4 — Budget estimate (rough, per person, KRW)
# ============================================================
wb_ = wb.create_sheet("Budget (rough)")
bh=["Item","KRW (pp)","ZAR (pp)","Notes"]; bwd=[40,14,12,45]
for i,(h,w) in enumerate(zip(bh,bwd),1):
    c=wb_.cell(1,i,h); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
    wb_.column_dimensions[get_column_letter(i)].width=w
wb_.freeze_panes="A2"
# (item, krw, note)
budget=[
    ("Everland ticket (Klook)",39000,"vs ~62k at gate"),
    ("Gyeongbokgung",3000,"Free in hanbok"),
    ("MMCA Seoul",4000,"Most national museums FREE"),
    ("Seokpajeong (+ Seoul Museum)",20000,"Fri"),
    ("N Seoul Tower (cable car + obs)",35000,"Optional"),
    ("COEX Aquarium",33000,"Optional"),
    ("Jeju rental car (3 days, per car)",180000,"SPLIT between travellers"),
    ("Jeju site fees (Manjanggul, Seongsan, folk, cliffs, falls)",43000,"Across both Jeju days"),
    ("Udo ferry (return)",10500,"East day"),
    ("Van Gogh Garden + Alive Museum",27000,"South day"),
    ("Songdo cable car (RT general)",17000,"Crystal floor ₩22k"),
    ("Haeundae Sky Capsule (per 2-seat capsule)",17500,"₩35k/capsule ÷ 2"),
    ("The Bay 101 yacht",20000,"Night cruise"),
    ("Baseball ticket",13000,"If Lotte home"),
    ("KTX Busan->Seoul (1 Aug)",59800,"+ AREX ₩4,750"),
    ("AREX Seoul->Incheon (express)",4750,"1 Aug"),
    ("Food (avg/day)",40000,"Markets cheaper"),
    ("T-money transit (per city/day)",10000,"Top up as needed"),
]
for ri,(item,krw,bnote) in enumerate(budget,2):
    zar = round(krw/RATE)
    for ci,v in enumerate([item, f"₩{krw:,}", f"R{zar:,}", bnote],1):
        c=wb_.cell(ri,ci,v); c.border=BORDER; c.alignment=WRAP
nr=len(budget)+3
note=wb_.cell(nr,1,"NB: Rough PER-PERSON estimates. Converted at R1 = ₩93 (June 2026). Flights & hotels excluded (already booked). Many top museums (National Museum, Folk Museum, War Memorial, Seoul Museum of History, Seoul Forest, Bukchon, Banpo fountain) are FREE. Jeju car cost is per car — divide by the number of travellers.")
wb_.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=4)
note.alignment=WRAP; note.font=Font(italic=True)
wb_.row_dimensions[nr].height=60

# ---- TOTAL TRIP ESTIMATE: Lean vs Comfortable (per person) ----
sr = nr + 2
title = wb_.cell(sr,1,"TOTAL TRIP ESTIMATE — per person (excl. flights & hotels, ~13 days on ground)")
wb_.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=4)
title.fill=HDR_FILL; title.font=HDR_FONT; title.alignment=Alignment(horizontal="left",vertical="center"); title.border=BORDER
sr += 1
sh=["Category","Lean (₩)","Comfortable (₩)","What's included / assumption"]
for ci,h in enumerate(sh,1):
    c=wb_.cell(sr,ci,h); c.fill=DAY_FILL; c.font=DAY_FONT; c.alignment=CENTER; c.border=BORDER
sr += 1
# (category, lean_krw, comf_krw, note)
scen=[
    ("Attractions & activities",160000,360000,"Lean=core sights only; Comf=all optionals (N Seoul Tower, Aquarium, Leeum, Folk Village, Van Gogh, Alive, Busan Tower, Seogwipo submarine)"),
    ("Food & drink",350000,650000,"~13 days; Lean ~₩27k/day (markets), Comf ~₩50k/day (restaurants, cafes, drinks)"),
    ("Local transport (metro/bus/taxi)",100000,200000,"Lean=mostly T-money; Comf=more taxis"),
    ("Intercity (KTX + AREX + airport transfers)",78000,105000,"KTX Busan→Seoul, AREX, ICN/Gimpo/Gimhae transfers"),
    ("Jeju rental car (your share)",75000,105000,"Per CAR ÷ 2 travellers; Comf=bigger car + full insurance"),
    ("Misc buffer (SIM, souvenirs, snacks)",50000,150000,"Padding for the unexpected"),
]
lean_total=0; comf_total=0
for cat,lean,comf,bnote in scen:
    lean_total+=lean; comf_total+=comf
    vals=[cat, f"₩{lean:,}", f"₩{comf:,}", bnote]
    for ci,v in enumerate(vals,1):
        c=wb_.cell(sr,ci,v); c.border=BORDER; c.alignment=WRAP
    wb_.row_dimensions[sr].height=42
    sr+=1
# KRW total row
for ci,v in enumerate(["TOTAL (₩)", f"₩{lean_total:,}", f"₩{comf_total:,}", ""],1):
    c=wb_.cell(sr,ci,v); c.border=BORDER; c.fill=WARN_FILL; c.font=Font(bold=True); c.alignment=WRAP
sr+=1
# ZAR total row
for ci,v in enumerate(["TOTAL (ZAR)", f"R{round(lean_total/RATE):,}", f"R{round(comf_total/RATE):,}", "≈ per person at R1=₩93"],1):
    c=wb_.cell(sr,ci,v); c.border=BORDER; c.fill=WARN_FILL; c.font=Font(bold=True,color="1F4E78"); c.alignment=WRAP

out = "/home/tevlen/Projects/SK26/TNT_South_Korea_Itinerary_2026.xlsx"
wb.save(out)
print("Saved", out)
print("Sheets:", wb.sheetnames)
